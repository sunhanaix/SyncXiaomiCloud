import os,sys,re,json
import openpyxl
from openpyxl.styles import Alignment,Border,Side
'''
给定一个从小米云上同步下来的通讯录信息，
生成一个excel格式的结果
'''
fname='contact.json'

def load_file(fname):
    ss=open(fname,'r',encoding='utf8').read()
    return json.loads(ss)

def convert_contact_to_xls(src,tgt):
    res=load_file(src)
    xls_file=tgt
    #print(json.dumps(res,indent=2,ensure_ascii=False))
    result=[]
    phone_dict={}
    for k in res:
        c=res[k]
        name=c['content']['displayName']
        if not c['content'].get('phoneNumbers') or not type( c['content']['phoneNumbers'])==list:
            continue
        phones=[]
        for p in  c['content']['phoneNumbers']:
            p['value']=p['value'].replace(' ','').replace('-','').replace('.','')
            p['value']=re.sub(r'^0086|^\+86','',p['value'])
            phones.append(p['value'])
        emails=[]
        if not c['content'].get('emails') or not type( c['content']['emails'])==list:
            emails=[]
        else:
            for e in c['content']['emails']:
                emails.append(e['value'])
        addresses=[]
        if not c['content'].get('addresses') or not type(c['content']['addresses'] )==list:
            addresses=[]
        else:
            for a in c['content']['addresses']:
                addresses.append(a['formatted'])
        company=[]
        if not c['content'].get('organizations') or not type(c['content']['organizations'] )==list:
            company=[]
        else:
            for c in c['content']['organizations']:
                for item in c:
                    company.append(c[item])
        result.append({'name':name,'phone':phones,'email':emails,'address':addresses,'company':company})
        for p in phones:
            phone_dict[p]=name
    open('res.json','w',encoding='utf8').write(json.dumps(result,indent=2,ensure_ascii=False))
    book = openpyxl.Workbook()
    border = Border(
        left=Side(
            border_style="thin",
            color="FF0000"
        ),
        right=Side(
            border_style="thin",
            color="FF0000"
        ),
        top=Side(
            border_style="thin",
            color="FF0000"
        ),
        bottom=Side(
            border_style="thin",
            color="FF0000"

        )
    )
    col_names = [chr(i) for i in range(65, 65 + 26)]  # 'A'...'Z'的列
    widths = {'name': 35, 'phone': 19, 'email': 29, 'address': 20, 'company': 19}
    col_widths = {}
    sheet=book.create_sheet(title='通讯录')
    row = 1
    col = 1
    id = 1
    for t in widths:
        sheet.cell(row, col, t).border=border
        col += 1
        id += 1
        col_widths[col - 1] = widths[t] # 要是列名字是预定义的，直接取数值
    row += 1
    for row_data in result:
        row_data['phone']='\n'.join(row_data['phone'])
        row_data['email'] = '\n'.join(row_data['email'])
        row_data['address'] = '\n'.join(row_data['address'])
        row_data['company'] = '\n'.join(row_data['company'])
        col = 1
        for k in row_data:
            sheet.cell(row, col, row_data[k]).alignment = Alignment(wrapText=True,vertical='center')
            sheet.cell(row,col).border=border
            col+=1
        row+=1
    for col in col_widths:  # 给每一列调整下列宽
        sheet.column_dimensions[col_names[col - 1]].width = col_widths[col]
    try:
        del book['Sheet']  # 会默认多建一个叫"Sheet"的表，如果存在就把它删掉
    except Exception as e:
        pass
    book.save(xls_file)
    phone_dict_json=os.path.splitext(tgt)[0]+'.json'
    open(phone_dict_json,'w',encoding='utf8').write(json.dumps(phone_dict,indent=2,ensure_ascii=False))
    return phone_dict_json

if __name__=='__main__':
    convert_contact_to_xls(src=fname,tgt='contacts.xlsx')
