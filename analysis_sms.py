import os,sys,re,json,time
import openpyxl
from openpyxl.styles import Alignment,Border,Side
import difflib
'''
给定一个从小米云上同步，并格式化好的sms.json文件
以及格式化好的通讯录电话为key的文件
输出短信记录，到一个excel表里面
'''

fname='sms.json'
contact_json_file='phone_dict.json'

def simRatio(s1, s2):  # 比较两个字符串的相似度，结果为0-1之间的数
    s = difflib.SequenceMatcher(None, s1, s2)
    return s.ratio()

def load_file(fname):
    ss=open(fname,'r',encoding='utf8').read()
    return json.loads(ss)

def get_name_by_phone(phone,phone_dict):
    #给定一个号码，先看直接命中了没有，命中了的话，就返回这个对应名字
    #如果没有命中，看它是不是其中某几个号码的子集，如果是子集，再这几个匹配到了号码里面，选一个相似度最大的一个号码返回
    if phone in phone_dict:
        return phone_dict[phone]
    res=None
    max_r=0
    for p in phone_dict:
        if p.find(phone)>-1:
            r=simRatio(p,phone)
            if r>max_r:
                max_r=r
                res=p
    if not res or max_r<0.8: #要是没有匹配到，或者匹配到的相似系数不够大，那么就返回空
        return ''
    return phone_dict[res]

def convert_sms_to_xls(src,tgt,contact=contact_json_file):
    sms=load_file(src)
    phone_dict=load_file(contact)
    xls_file=tgt
    res=[]
    for s in sms:
        txt=s['entry']['snippet']
        ts=s['entry']['localTime']/1000
        date_time=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts) )
        sender=s['entry']['recipients']
        sender=re.sub(r'^\+86|^0086','',sender)
        name=get_name_by_phone(sender,phone_dict)
        res.append({'name':name,'sender':sender,'time':date_time,'txt':txt,})
    open('sms_res.json','w',encoding='utf8').write(json.dumps(res,indent=2,ensure_ascii=False))
    book = openpyxl.Workbook()
    border = Border(
        left=Side(
            border_style="thin",
            color="FF0000"
        ),
        right=Side(
            border_style="thin",
            color="FF0000"
        ),
        top=Side(
            border_style="thin",
            color="FF0000"
        ),
        bottom=Side(
            border_style="thin",
            color="FF0000"

        )
    )
    col_names = [chr(i) for i in range(65, 65 + 26)]  # 'A'...'Z'的列
    widths = {'name': 20, 'sender': 22, 'time':21,'txt':110}
    col_widths = {}
    sheet = book.create_sheet(title='短信')
    row = 1
    col = 1
    id = 1
    for t in widths:
        sheet.cell(row, col, t).border = border
        col += 1
        id += 1
        col_widths[col - 1] = widths[t]  # 要是列名字是预定义的，直接取数值
    row += 1
    for row_data in res:
        col = 1
        for k in row_data:
            try:
                sheet.cell(row, col).alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row, col).border = border
                sheet.cell(row, col).value = row_data[k]
            except Exception as e:
                print(e)
                print(row_data[k])
            col += 1
        row += 1
    for col in col_widths:  # 给每一列调整下列宽
        sheet.column_dimensions[col_names[col - 1]].width = col_widths[col]
    try:
        del book['Sheet']  # 会默认多建一个叫"Sheet"的表，如果存在就把它删掉
    except Exception as e:
        pass
    book.save(xls_file)
    return res
if __name__=='__main__':
    convert_sms_to_xls(src=fname,tgt='sms.xlsx')
